# -*- coding: utf-8 -*-
# Define your item pipelines here
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import os
import xlrd
import shutil
import zipfile
import openpyxl
import tempfile
import const as ct
import dspider.poster as poster
from dspider import items
from scrapy import Request
from scrapy.exceptions import DropItem
from tempfile import TemporaryDirectory
from scrapy.pipelines.files import FilesPipeline

post_router = {
    items.MyDownloadItem:poster.MyDownloadItemPoster,
    items.PlateValuationItem:poster.PlateValuationPoster,
    items.HkexTradeTopTenItem:poster.HkexTradeTopTenItemPoster,
    items.HkexTradeOverviewItem:poster.HkexTradeOverviewPoster,
    items.SPledgeSituationItem:poster.SPledgeSituationItemPoster,
    items.InvestorSituationItem:poster.InvestorSituationItemPoster,
    items.MonthInvestorSituationItem:poster.MonthInvestorSituationItemPoster,
}

class DspiderPipeline(object):
    def process_item(self, item, spider):
        obj = post_router[item.__class__](item)
        if obj.check():
            obj.store()
        return item

class PlateValuationHandlePipeline(object):
    def process_item(self, item, spider):
        fnames = item['file_name']
        fdir = ct.PLATE_VALUATION_PATH
        for fname in fnames:
            with zipfile.ZipFile(os.path.join(fdir, fname)) as f:
                with TemporaryDirectory() as dirpath:
                    for mfile in f.namelist():
                        f.extract(mfile, dirpath)
                        item_list = self.create_item(dirpath, mfile)
                        for mitem in item_list: self.store(mitem)
        return item
       
    def store(self, item):
        obj = post_router[item.__class__](item)
        if obj.check(): obj.store()
        
    def create_item(self, fdir, fname):
        cdict = ct.PLATE_DICT
        fpath = os.path.join(fdir, fname)
        cdate = "%s-%s-%s" % (fname[2:6], fname[6:8], fname[8:10])
        try:
            wb = xlrd.open_workbook(fpath, encoding_override="cp1252")
        except Exception as e:
            return list()
        static_pe_sheet = wb.sheet_by_name('板块静态市盈率')
        rolling_pe_sheet = wb.sheet_by_name('板块滚动市盈率')
        pb_sheet = wb.sheet_by_name('板块市净率')
        dividend_yield_sheet = wb.sheet_by_name('板块股息率')
        item_list = list()
        for row in range(1, static_pe_sheet.nrows):
            item = items.PlateValuationItem()
            item['date'] = cdate
            name = static_pe_sheet.cell(row, 0).value
            item['name'] = name
            item['code'] = ct.PLATE_DICT[name]
            item['pe'] = item.convert(static_pe_sheet.cell(row, 1).value)
            item['ttm'] = item.convert(rolling_pe_sheet.cell(row, 1).value)
            item['pb'] = item.convert(pb_sheet.cell(row, 1).value)
            item['dividend'] = item.convert(dividend_yield_sheet.cell(row, 1).value)
            item_list.append(item)
        return item_list

class PlateValuationDownloadPipeline(FilesPipeline):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def file_path(self, request, response = None, info=None):
        return request.meta['item']['file_name'].strip('";')

    def get_media_requests(self, item, info):
        # FilesPepeline 根据file_urls指定的url进行爬取，该方法为每个url生成一个Request后 Return
        for file_url in item['file_urls']:
            yield Request(file_url, meta={'item': item})

    def item_completed(self, results, item, info):
        file_urls = [x['path'] for ok, x in results if ok]
        if not file_urls: raise DropItem("Item contains no files")
        item['file_name'] = file_urls
        return item

class SPledgePipline(FilesPipeline):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def get_media_requests(self, item, info):
        # FilesPepeline 根据file_urls指定的url进行爬取，该方法为每个url生成一个Request后 Return
        for file_url in item['file_urls']:
            yield Request(file_url, meta={'item': item})

    def item_completed(self, results, item, info):
        file_urls = [x['path'] for ok, x in results if ok]
        if not file_urls: raise DropItem("Item contains no files")
        item['file_name'] = file_urls
        return item

    def file_path(self, request, response = None, info=None):
        return request.meta['item']['file_name'].replace('gpzyhgmx_', '')
